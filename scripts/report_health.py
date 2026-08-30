#!/usr/bin/env python3
"""
report_health.py — 건강검진 JSON → CSV → 엑셀 리포트 (f_xlsx 계열)

ADB ↔ 엑셀 연결의 "렌더부". 원본은 phone-health.sh 가 매일 쌓는 JSON,
이걸 CSV(진실의 원본)로 모으고, f_xlsx 스타일로 손으로 만든 것처럼 렌더링한다.

데이터 수집은 두 갈래:
  - 지금: phone-health.sh (proot, /proc 기반) → _notebook/health/*.json
  - 다음: adb_collect.sh (WSL, dumpsys 기반) → 같은 CSV 에 append

사용법:
  python3 scripts/report_health.py
  → data/health_log.csv + output/health_report.xlsx
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd

from render_xlsx import BORDER, GRAY, LIGHT, NAVY, TEAL, _num_fmt  # 스타일 재사용
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).resolve().parent.parent
HEALTH_DIR = BASE / "_notebook" / "health"
CSV_PATH = BASE / "data" / "health_log.csv"
OUT_PATH = BASE / "output" / "health_report.xlsx"

COLS = ["date", "grade", "pass", "warn", "fail", "total", "mem_avail_mb", "swap_pct"]


def json_to_rows() -> list[dict]:
    """health/*.json (phone-health.sh 출력) → 시간순 행 목록."""
    rows = []
    for f in sorted(HEALTH_DIR.glob("2026-*.json")):
        raw = f.read_text(encoding="utf-8")
        # phone-health.sh 버그: "http_status": 000000 (선행 0 숫자 = JSON 위반) → 문자열로 교정
        raw = re.sub(r'(:\s*)0{2,}(\d+)', r'\1"\2"', raw)
        try:
            d = json.loads(raw)
        except json.JSONDecodeError:
            continue
        r = d.get("results", {})
        mem = d.get("memory", {})
        rows.append({
            "date": d.get("timestamp", "")[:10],
            "grade": d.get("grade", "?"),
            "pass": r.get("pass"), "warn": r.get("warn"),
            "fail": r.get("fail"), "total": r.get("total"),
            "mem_avail_mb": mem.get("available_mb"),
            "swap_pct": mem.get("swap_pct"),
        })
    return rows


def render(df: pd.DataFrame, out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "건강리포트"

    n = len(df)
    head_row = 5
    first = head_row + 1
    last = head_row + n

    ws["A1"] = "폰 건강 검진 리포트"
    ws["A1"].font = Font(bold=True, size=16, color=NAVY)
    ws["A2"] = f"자동 생성 · {datetime.now():%Y-%m-%d} · 원본 phone-health.sh"
    ws["A2"].font = Font(size=9, color=GRAY)
    ws.merge_cells("A1:H1"); ws.merge_cells("A2:H2")

    # 테이블 헤더
    label_map = {"date": "날짜", "grade": "등급", "pass": "통과", "warn": "경고",
                 "fail": "실패", "total": "항목", "mem_avail_mb": "가용메모리(MB)", "swap_pct": "스왑(%)"}
    for j, col in enumerate(COLS, start=1):
        c = ws.cell(row=head_row, column=j, value=label_map[col])
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    # 데이터 값
    for i, (_, row) in enumerate(df.iterrows()):
        r = first + i
        for j, col in enumerate(COLS, start=1):
            v = row[col]
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
            if col == "date":
                c.alignment = Alignment(horizontal="center")
            elif col == "grade":
                c.alignment = Alignment(horizontal="center")
                c.font = Font(bold=True, color=NAVY)
            else:
                c.number_format = _num_fmt(col)
                c.alignment = Alignment(horizontal="right")
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=LIGHT)

    # 조건부 서식 — 실패 수 (fail 컬럼) 컬러 스케일
    ws.conditional_formatting.add(
        f"E{first}:E{last}",
        ColorScaleRule(start_type="min", start_color="D1FAE5",
                       end_type="max", end_color="FEE2E2"),
    )

    # 차트 — 가용 메모리 추이
    chart = LineChart(); chart.title = "가용 메모리 추이"; chart.style = 13
    chart.y_axis.title = "MB"; chart.height = 8; chart.width = 16
    data = Reference(ws, min_col=7, min_row=head_row, max_row=last)
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    ws.add_chart(chart, "J2")

    for j, col in enumerate(COLS, start=1):
        ws.column_dimensions[chr(64 + j)].width = {"date": 12, "grade": 7, "pass": 7,
                                                   "warn": 7, "fail": 7, "total": 7,
                                                   "mem_avail_mb": 16, "swap_pct": 10}.get(col, 10)
    ws.freeze_panes = f"A{first}"

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"✅ {out} 생성 — 검진 {n}회")


def main() -> int:
    rows = json_to_rows()
    if not rows:
        print("❌ health JSON 없음", file=__import__("sys").stderr)
        return 1
    df = pd.DataFrame(rows)[COLS]
    CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(CSV_PATH, index=False)
    print(f"✅ {CSV_PATH} 갱신 — {len(rows)}행")
    render(df, OUT_PATH)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
