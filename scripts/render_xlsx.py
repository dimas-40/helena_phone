#!/usr/bin/env python3
"""
f_xlsx — 로데이터(CSV) → 엑셀 리포트 렌더러 (렌더 모드 패밀리: f_html/f_xlsx/f_png/f_pptx)

엑셀은 DB가 아니라 **출력물**이다.
  - 데이터(CSV) = 진실의 원본 (텍스트라 git diff 가능)
  - 연산(pandas) = 코드 (읽힘·버전관리)
  - .xlsx      = 리포트 (매번 재생성, .gitignore 대상)
  - 요약셀 몇 개만 "수식"으로, 나머진 "값" — 받는 사람이 숫자 바꿔봐도 요약이 살아있게

설계 원칙: **서식이 8할.** 연산은 pandas 3줄, 열너비·숫자포맷·조건부서식·차트가 코드의 대부분.
템플릿 한 번 잘 만들면 그다음부터 공짜.

사용법:
  python3 scripts/render_xlsx.py data/weekly_metrics.csv
  → output/weekly_report.xlsx
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 팔레트 ──────────────────────────────────────────────
NAVY = "1F3864"      # 제목·테이블 헤더
TEAL = "14B8A6"      # 섹션 헤더 (기존 accent 계열)
LIGHT = "F2F5F9"     # 교차 행
GRAY = "8A94A6"      # 부제·메모

THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _num_fmt(col: str) -> str:
    return "+#,##0;-#,##0" if col in ("followers_delta",) else "#,##0"


def render_xlsx(df: pd.DataFrame, out_path: Path, title: str, source: str) -> None:
    """df → 손으로 만든 것처럼 보이는 xlsx 리포트."""
    wb = Workbook()
    ws = wb.active
    ws.title = "주간리포트"

    n = len(df)                      # 데이터 행 수
    head_row = 11                    # 테이블 헤더 행
    first_data = head_row + 1        # 첫 데이터 행
    last_data = head_row + n         # 마지막 데이터 행

    # ── 제목 / 부제 ────────────────────────────────────
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color=NAVY)
    ws["A2"] = f"자동 생성 · {datetime.now():%Y-%m-%d} · 원본 {source}"
    ws["A2"].font = Font(size=9, color=GRAY)
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    ws.merge_cells(f"A2:{get_column_letter(len(df.columns))}2")

    # ── 요약 블록 (수식 몇 개 = "사람이 만든 것처럼") ──
    def section(row, text):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=11, color=TEAL)

    section(4, "요약")
    summary = [
        ("총 조회수",      f"=SUM(C{first_data}:C{last_data})"),
        ("평균 주간 조회수", f"=ROUND(AVERAGE(C{first_data}:C{last_data}),0)"),
        ("총 팔로워 증가",  f"=SUM(E{first_data}:E{last_data})"),
        ("최고 조회 주",   f"=INDEX(A{first_data}:A{last_data},MATCH(MAX(C{first_data}:C{last_data}),C{first_data}:C{last_data},0))"),
    ]
    for i, (label, formula) in enumerate(summary):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        v = ws.cell(row=r, column=2, value=formula)
        v.font = Font(bold=True, size=12, color=NAVY)
        v.number_format = "#,##0"
        v.fill = PatternFill("solid", fgColor="EAF7F5")
        v.border = BORDER

    # ── 테이블 헤더 ─────────────────────────────────────
    section(9, "주간 상세")
    cols = list(df.columns)
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=head_row, column=j, value=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    # ── 데이터 값 (수식 없음 — 값만) ──────────────────
    for i, (_, row) in enumerate(df.iterrows()):
        r = first_data + i
        for j, col in enumerate(cols, start=1):
            c = ws.cell(row=r, column=j, value=row[col])
            c.border = BORDER
            if j == 1:
                c.alignment = Alignment(horizontal="center")
                c.font = Font(bold=True)
            else:
                c.number_format = _num_fmt(col)
                c.alignment = Alignment(horizontal="right")
            if i % 2 == 1:  # 교차 행
                c.fill = PatternFill("solid", fgColor=LIGHT)

    # ── 조건부 서식 (조회수 컬러 스케일) ────────────────
    views_col = get_column_letter(cols.index("views") + 1)
    ws.conditional_formatting.add(
        f"{views_col}{first_data}:{views_col}{last_data}",
        ColorScaleRule(
            start_type="min", start_color="FEE2E2",
            mid_type="percentile", mid_value=50, mid_color="FEF3C7",
            end_type="max", end_color="D1FAE5",
        ),
    )

    # ── 차트 (조회수 추이) ──────────────────────────────
    chart = LineChart()
    chart.title = "주간 조회수 추이"
    chart.style = 13
    chart.y_axis.title = "조회수"
    chart.height = 8
    chart.width = 16
    data = Reference(ws, min_col=cols.index("views") + 1, min_row=head_row, max_row=last_data)
    cats = Reference(ws, min_col=1, min_row=first_data, max_row=last_data)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"G4")

    # ── 열너비 (서식의 절반) ───────────────────────────
    widths = {"week": 10, "posts": 8, "views": 12, "likes": 10, "followers_delta": 14}
    for j, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 12)

    ws.freeze_panes = f"A{first_data}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"✅ {out_path} 생성 — {n}행 · 요약수식 {len(summary)}개 · 차트 1개")


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    src = Path(sys.argv[1])
    if not src.exists():
        print(f"❌ 없음: {src}", file=sys.stderr)
        return 1
    df = pd.read_csv(src)
    out = Path("output") / f"{src.stem}_report.xlsx"
    render_xlsx(df, out, title="주간 콘텐츠 성과 리포트", source=src.name)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
